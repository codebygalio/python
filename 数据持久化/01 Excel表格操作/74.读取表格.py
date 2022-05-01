import openpyxl

# 根据文件路径读取表格文件
workbook = openpyxl.load_workbook('实例.xlsx')

# workbook.sheetnames  获取所有表的名字
print(workbook.sheetnames)

sheet2 = workbook['表2']
print(sheet2)

# 在读取数据之前, 需要获取当前表行列范围
print(sheet2.max_row)  # 获取最大行
print(sheet2.max_column)  # 获取最大列

# 取第一行数据
# for i in range(1, sheet2.max_column + 1):
#     print(sheet2.cell(row=1, column=i).value)

# 取第一列数据
# for j in range(1, sheet2.max_row + 1):
#     print(sheet2.cell(row=j, column=1).value)

for i in range(1, sheet2.max_column + 1):
    for j in range(1, sheet2.max_row + 1):
        print(sheet2.cell(row=i, column=j).value)



