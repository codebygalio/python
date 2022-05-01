import openpyxl  # 导入操作表格文件的模块

# 1.创建工作簿对象
work_book = openpyxl.Workbook()

# 2. 很具工作簿对象创建表
sheet1 = work_book.create_sheet('表1')

# 3.数据写入表格
# sheet1['B3'] = 'B3'
# sheet1['B10'] = 'B10'

sheet1.cell(row=1, column=1).value = '111111'
sheet1.cell(row=2, column=2).value = '222222'

# 爬虫数据采集, 可以采集若干个字段
# sheet1.append(list1)  整行添加数据到表格中, 括号内部需要传有序列的数据容器
list1 = [1, 2, 3, 4, 5]
list2 = (1, 2, 3, 4, "5")
sheet1.append(list1)
sheet1.append(list2)

# 4. 保存表格文件
work_book.save('实例.xlsx')
