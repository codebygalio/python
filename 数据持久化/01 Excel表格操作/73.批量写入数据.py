
import openpyxl


# 创建工作簿对象
work = openpyxl.Workbook()
# 创建表
sheet = work.create_sheet('表2')
# 写入数据
for i in range(1, 10):
    for j in range(1, i + 1):
        print(f'{j} * {i} = {i * j}', end='\t')
        sheet.cell(row=i, column=j).value = f'{j} * {i} = {i * j}'
    print()


# 保存数据
work.save('实例.xlsx')





